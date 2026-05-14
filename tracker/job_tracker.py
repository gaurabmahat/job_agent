# tracker/job_tracker.py
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TRACKER_PATH = "job_tracker.xlsx"

# -- Column definitions ------------------------------------------------
COLUMNS = ["Company", "Job Role", "Job Description", "Applied Date"]

# Column widths (in Excel units)
COLUMN_WIDTHS = {
    "A": 20,   # Company
    "B": 35,   # Job Role
    "C": 80,   # Job Description
    "D": 15,   # Applied Date
}


def _style_header_row(sheet):
    """Applies header styling - dark background, white bold text."""
    header_fill = PatternFill("solid", start_color="2D3748")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for col_num, col_name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False
        )


def _style_data_row(sheet, row_num: int):
    """Applies alternating row colour and text wrapping to a data row."""
    # Alternate between white and very light grey
    fill_color = "FFFFFF" if row_num % 2 == 0 else "F7FAFC"
    row_fill = PatternFill("solid", start_color=fill_color)
    row_font = Font(name="Arial", size=10)

    for col_num in range(1, len(COLUMNS) + 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.fill = row_fill
        cell.font = row_font
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )


def _apply_column_widths(sheet):
    for col_letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[col_letter].width = width


def _create_new_tracker() -> Workbook:
    """Creates a fresh tracker workbook with headers and styling."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Applications"

    # Freeze the header row so it stays visible when scrolling
    sheet.freeze_panes = "A2"

    _style_header_row(sheet)
    _apply_column_widths(sheet)

    # Set header row height
    sheet.row_dimensions[1].height = 25

    return wb

def save_to_tracker(
    company: str,
    job_role: str,
    job_description: str,
    tracker_path: str = TRACKER_PATH
) -> bool:
    """
    Appends a job application entry to the Excel tracker.
    Creates the file if it doesn't exist yet.
    Returns True on success, False on failure.
    """

    today = datetime.now().strftime("%Y-%m-%d")

    try:
        # -- Load existing or create new ------------------------------------------
        if os.path.exists(tracker_path):
            wb = load_workbook(tracker_path)
            sheet = wb.active
        else:
            print(f"[TRACKER] No tracker found - creating new file: {tracker_path}")
            wb = _create_new_tracker()
            sheet = wb.active

        # -- Find the next empty row ------------------------------------------------
        next_row = sheet.max_row + 1

        # -- Write the data ---------------------------------------------------------
        trimmed_description = job_description[:1000].strip()
        if len(job_description) > 1000:
            trimmed_description += "... [trimmed]"

        row_data = [company, job_role, trimmed_description, today]

        for col_num, value in enumerate(row_data, start=1):
            sheet.cell(row=next_row, column=col_num, value=value)

        # -- Style the new row -------------------------------------------------------
        _style_data_row(sheet, next_row)
        sheet.row_dimensions[next_row].height = 60

        # -- Save with retry loop -----------------------------------------------------
        while True:
            try:
                wb.save(tracker_path)
                print(f"[TRACKER] Saved to tracker: {tracker_path}")
                return True

            except PermissionError:
                print(f"\n[ERROR] Cannot save - {tracker_path} is open in Excel.")
                print("[FIX]   Close the file in Excel, then press Enter to try again...")
                input()  # Wait for user to close Excel

    except Exception as e:
        print(f"\n[ERROR] Tracker save failed unexpectedly: {e}")
        return False
    